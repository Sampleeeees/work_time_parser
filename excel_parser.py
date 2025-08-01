"""
Excel parser for removing unused columns and generating financial or project-based reports.

Supports grouping by full month or split-half (1–15, 16–end).
"""
import calendar
from collections import defaultdict
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, DEFAULT_FONT, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_NUMBER
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from enums import ReportGroupingType


class ExcelParser:
    """Excel parser class for generating reports."""

    UNUSED_COLUMNS: list = [
        "A", "C", "D",  # ID, Date/Time, End date/time
        "F", "H", "I",  # Who, Project Category, Company
        "J", "L", "M",  # Task list, Parent task, is sub-task
        "N", "O", "P",  # Is billable?, Invoice number, Hours
        "Q", "S", "T",  # Minutes, Estimated time, Estimated(hours),
        "U", "V", "W",  # Estimated(minutes), Tags, Task tags,
        "X", "Y", "Z",  # First name, Last name, User ID
    ]

    def __init__(self, workbook_path: str) -> None:
        """Initialize ExcelParser with the path to the workbook."""
        self.workbook_path: str = workbook_path
        DEFAULT_FONT.name = "Ubuntu"

    @staticmethod
    def generate_report_filename(report_type: str, date: datetime) -> str:
        """Generate a report filename like 'financial_august_report.xlsx'"""
        mount_name: str = date.strftime("%B").lower()
        return f"{report_type}_{mount_name}_report.xlsx"

    @staticmethod
    def _set_column_widths(ws: Worksheet, widths: list[int]) -> None:
        """Set column widths for a given worksheet."""
        for col_index, width in enumerate(widths, start=1):
            col_letter = ws.cell(row=1, column=col_index).column_letter
            ws.column_dimensions[col_letter].width = width

    @staticmethod
    def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
        """Write a bold header row with gray background."""
        header_font = Font(name=DEFAULT_FONT.name, bold=True)
        header_fill = PatternFill(fill_type="solid",start_color="BDBDBD", end_color="BDBDBD")
        for col_index, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_index, value=value)
            cell.font = header_font
            cell.fill = header_fill

    @staticmethod
    def _merge_title_row(ws: Worksheet, row: int, title: str) -> None:
        """Merge 8 columns for a title row and apply styling."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(name=DEFAULT_FONT.name, bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color="CFE2F3", end_color="CFE2F3")

    @staticmethod
    def _parse_date(raw_date) -> datetime | None:
        """Convert raw date value to datetime object safely."""
        if isinstance(raw_date, datetime):
            return raw_date
        if isinstance(raw_date, str):
            try:
                return datetime.fromisoformat(raw_date)
            except ValueError:
                return None
        return None

    @staticmethod
    def build_period_titles(date: datetime) -> dict[str, str]:
        month = date.month
        year = date.year
        last_day = calendar.monthrange(year, month)[1]

        return {
            "01–15": f"Period: 01.{month:02}.{year} – 15.{month:02}.{year}",
            "16–31": f"Period: 16.{month:02}.{year} – {last_day}.{month:02}.{year}"
        }

    @staticmethod
    def _write_bottom_totals(
        ws: Worksheet,
        row_cursor: int,
        last_data_row: int,
        rate: int,
        exchange_rate: float,
        rate_row: int,
        exch_row: int
    ) -> None:
        """Write total hours, total UAH, and total USD summary at the bottom of the worksheet."""
        # insert rate and exchange at bottom
        ws.cell(row=rate_row, column=6, value="Рейт").font = Font(name=DEFAULT_FONT.name, bold=True)
        ws.cell(row=exch_row, column=6, value="Курс").font = Font(name=DEFAULT_FONT.name, bold=True)

        rate_cell = ws.cell(row=rate_row, column=7, value=rate)
        rate_cell.font = Font(name=DEFAULT_FONT.name, bold=True)
        rate_cell.number_format = FORMAT_NUMBER_00

        exchange_cell = ws.cell(row=exch_row, column=7, value=exchange_rate)
        exchange_cell.font = Font(name=DEFAULT_FONT.name, bold=True)
        exchange_cell.number_format = FORMAT_NUMBER_00

        # totals
        ws.cell(row=row_cursor, column=2, value="ИТОГО часов").font = Font(name=DEFAULT_FONT.name, bold=True)
        sum_hours_cell = ws.cell(row=row_cursor, column=4, value=f"=SUM(D2:D{last_data_row})")
        sum_hours_cell.number_format = FORMAT_NUMBER_00
        ws.cell(row=row_cursor, column=5, value="часов").alignment = Alignment(horizontal="right")

        ws.cell(row=row_cursor + 1, column=2, value="ИТОГО к оплате").font = Font(name=DEFAULT_FONT.name, bold=True)
        sum_uah_cell = ws.cell(row=row_cursor + 1, column=4, value=f"=SUM(G2:G{last_data_row})")
        sum_uah_cell.number_format = FORMAT_NUMBER_00
        ws.cell(row=row_cursor + 1, column=5, value="UAH").alignment = Alignment(horizontal="right")

        ws.cell(row=row_cursor + 2, column=2, value="ИТОГО в долларах").font = Font(name=DEFAULT_FONT.name, bold=True)
        sum_usd_cell = ws.cell(
            row=row_cursor + 2,
            column=4,
            value=f"=SUM(G2:G{last_data_row})/{exchange_rate}"
        )
        sum_usd_cell.number_format = FORMAT_NUMBER_00
        ws.cell(row=row_cursor + 2, column=5, value="$").alignment = Alignment(horizontal="right")

        summary_fill = PatternFill(fill_type="solid", start_color="E8F0FE", end_color="E8F0FE")
        for r in range(row_cursor, row_cursor + 3):
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = summary_fill

    def remove_unused_columns(self, is_project_report: bool = False) -> None:
        """Remove unused columns."""
        wb: Workbook = load_workbook(self.workbook_path)
        ws: Worksheet = wb.active

        # if its project report then doesn't delete S column (Estimated time in hours)
        if is_project_report:
            self.UNUSED_COLUMNS.pop(14)

        indices = [column_index_from_string(col) for col in self.UNUSED_COLUMNS]
        for col_index in sorted(indices, reverse=True):
            ws.delete_cols(col_index)

        wb.save(self.workbook_path)

    def _write_financial_row(
        self,
        ws: Worksheet,
        row_cursor: int,
        row_data: tuple,
        rate_cell_ref: str,
        exchange_cell_ref: str
    ) -> int:
        """Write a single financial row with formulas and formating."""
        date_val = self._parse_date(row_data[0])
        desc = row_data[2] or ""
        task_name = row_data[3] or ""
        hours = float(row_data[4] or 0.0)
        task_id = row_data[5]
        teamwork_link = f"https://avada.teamwork.com/#tasks/{task_id}"

        ws.cell(row=row_cursor, column=1, value=task_id)
        cell = ws.cell(row=row_cursor, column=2, value=teamwork_link)
        cell.hyperlink = teamwork_link
        cell.style = "Hyperlink"
        ws.cell(row=row_cursor, column=3, value=desc)

        ws.cell(row=row_cursor, column=4, value=hours).number_format = FORMAT_NUMBER_00
        ws.cell(row=row_cursor, column=5, value=f"={rate_cell_ref}").number_format = FORMAT_NUMBER_00
        ws.cell(row=row_cursor, column=6, value=f"={exchange_cell_ref}").number_format = FORMAT_NUMBER_00

        amount_formula = f"=D{row_cursor}*E{row_cursor}*F{row_cursor}"
        ws.cell(row=row_cursor, column=7, value=amount_formula).number_format = FORMAT_NUMBER_00
        ws.cell(row=row_cursor, column=8, value=date_val.strftime("%d.%m.%Y") if date_val else "")

        return row_cursor + 1

    def generate_financial_report(self, rate: int, exchange_rate: float) -> str:
        """Generate a personal financial report grouped by project with total hours and costs."""
        self.remove_unused_columns()

        wb: Workbook = load_workbook(self.workbook_path)
        ws: Worksheet = wb.active

        # get header and rows data
        header, *data = ws.iter_rows(values_only=True)
        data.sort(key=lambda row: row[0])

        grouped = defaultdict(list)
        for row in data:
            grouped[row[1]].append(row)

        # precompute total rows to know where to place rate/exchange
        total_data_rows: int = sum(len(rows) + 4 for rows in grouped.values())

        rate_row = total_data_rows + 2
        exchange_row = rate_row + 1

        rate_cell_ref = f"$G${rate_row}"
        exchange_cell_ref = f"$G${exchange_row}"

        new_wb: Workbook = Workbook()
        new_ws: Worksheet = new_wb.active
        self._set_column_widths(new_ws, [15, 45, 55, 15, 15, 15, 20, 15])

        headers: list = [
            "Проект", "Ссылка на TeamWork", "Описание", "Кол-во часов",
            "Рейт/час [$]", "Курс [$]", "Стоимость (ГРН)", "Дата"
        ]

        self._write_header_row(new_ws, headers)

        row_cursor: int = 2

        for project, rows in grouped.items():
            self._merge_title_row(new_ws, row_cursor, project)
            row_cursor += 1

            for row in rows:
                row_cursor = self._write_financial_row(new_ws, row_cursor, row, rate_cell_ref, exchange_cell_ref)

            row_cursor += 3

        last_data_row = row_cursor - 4
        row_cursor += 1

        self._write_bottom_totals(new_ws, row_cursor, last_data_row, rate, exchange_rate, rate_row, exchange_row)

        filename: str = self.generate_report_filename("financial", data[0][0])

        new_wb.save(filename)

        return filename

    def generate_project_report(self, group_type: ReportGroupingType) -> str:
        """Generate a project report."""
        self.remove_unused_columns(is_project_report=True)

        wb: Workbook = load_workbook(self.workbook_path)
        ws: Worksheet = wb.active

        header, *data = ws.iter_rows(values_only=True)
        data.sort(key=lambda row: row[0])

        grouped = defaultdict(list)
        for row in data:
            grouped[row[1]].append(row)

        new_wb = Workbook()
        new_ws = new_wb.active
        self._set_column_widths(new_ws, [55, 50, 15, 15])

        self._write_header_row(new_ws, [
            "Task name", "Link to TeamWork", "Estimated time", "Total count time"
        ])

        row_cursor = 2

        if group_type == ReportGroupingType.SPLIT_HALF:
            filename = self.generate_report_filename(f"{group_type.value}_project", data[0][0])
            grouped_by_period = {
                "01–15": defaultdict(list),
                "16–31": defaultdict(list),
            }
            for row in data:
                date_val = self._parse_date(row[0])
                if not date_val:
                    continue
                key = "01–15" if date_val.day <= 15 else "16–31"
                project = row[1]
                grouped_by_period[key][project].append(row)

            period_titles = self.build_period_titles(date=data[0][0])

            for period_key in ["01–15", "16–31"]:
                # Period block title
                new_ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=4)
                period_cell = new_ws.cell(row=row_cursor, column=1, value=period_titles[period_key])
                period_cell.font = Font(name=DEFAULT_FONT.name, bold=True, italic=True)
                period_cell.fill = PatternFill(fill_type="solid", start_color="E0E0E0", end_color="E0E0E0")
                row_cursor += 1

                for project, rows in grouped_by_period[period_key].items():
                    new_ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=4)
                    cell = new_ws.cell(row=row_cursor, column=1, value=project)
                    cell.font = Font(name=DEFAULT_FONT.name, bold=True)
                    cell.fill = PatternFill(fill_type="solid", start_color="CFE2F3", end_color="CFE2F3")
                    row_cursor += 1

                    task_map = defaultdict(lambda: {"task_name": "", "hours": 0.0, "estimated": None})
                    for row in rows:
                        task_id = row[6]
                        task_name = row[3] or ""
                        hrs = float(row[4] or 0.0)
                        estimated = float(row[5] or 0.0)

                        task_map[task_id]["task_name"] = task_name
                        task_map[task_id]["hours"] += hrs
                        if task_map[task_id]["estimated"] is None and estimated is not None:
                            task_map[task_id]["estimated"] = estimated

                    for task_id, task in task_map.items():
                        new_ws.cell(row=row_cursor, column=1, value=task["task_name"])
                        url = f"https://avada.teamwork.com/#tasks/{task_id}"
                        cell = new_ws.cell(row=row_cursor, column=2, value=url)
                        cell.hyperlink = url
                        cell.style = "Hyperlink"
                        new_ws.cell(
                            row=row_cursor, column=3, value=task["estimated"] if task["estimated"] else None
                        ).number_format = FORMAT_NUMBER
                        new_ws.cell(row=row_cursor, column=4, value=task["hours"]).number_format = FORMAT_NUMBER_00
                        row_cursor += 1

                    row_cursor += 2  # space between projects

                row_cursor += 2  # space between periods
        else:
            filename = self.generate_report_filename(f"{group_type.value}_project", data[0][0])
            for project, rows in grouped.items():
                new_ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=4)
                cell = new_ws.cell(row=row_cursor, column=1, value=project)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", start_color="CFE2F3", end_color="CFE2F3")
                row_cursor += 1

                tasks = defaultdict(lambda: {"task_name": "", "hours": 0.0, "estimated": None})
                for row in rows:
                    task_id = row[6]
                    name = row[3] or ""
                    hrs = float(row[4] or 0.0)
                    estimated = float(row[5] or 0.0)
                    tasks[task_id]["task_name"] = name
                    tasks[task_id]["hours"] += hrs
                    if tasks[task_id]["estimated"] is None and estimated is not None:
                        tasks[task_id]["estimated"] = estimated

                for task_id, task in tasks.items():
                    new_ws.cell(row=row_cursor, column=1, value=task["task_name"])
                    url = f"https://avada.teamwork.com/#tasks/{task_id}"
                    cell = new_ws.cell(row=row_cursor, column=2, value=url)
                    cell.hyperlink = url
                    cell.style = "Hyperlink"
                    new_ws.cell(
                        row=row_cursor, column=3, value=task["estimated"] if task["estimated"] else None
                    ).number_format = FORMAT_NUMBER
                    new_ws.cell(row=row_cursor, column=4, value=task["hours"]).number_format = FORMAT_NUMBER_00
                    row_cursor += 1
                row_cursor += 2

        new_wb.save(filename)
        return filename
